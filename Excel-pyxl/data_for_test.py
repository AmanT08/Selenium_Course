import openpyxl


class DataForTest:

    @staticmethod
    def dataa(check_for_data):
        book = openpyxl.load_workbook("C:\\Users\\hp\\Documents\\Book1.xlsx")
        sheet = book.active

        dict = {}
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == check_for_data:
                for j in range(2, sheet.max_column + 1):
                    dict[sheet.cell(row=1, column=j).value]= sheet.cell(row=i, column=j).value
        return [dict]



