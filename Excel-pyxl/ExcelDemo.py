import openpyxl

book = openpyxl.load_workbook("C:\\Users\\hp\\Documents\\Book1.xlsx")
sheet = book.active
cells = sheet.cell(row=1,column=1)
print(cells.value)

# to write in cell
sheet.cell(row=6,column=1).value = "t5"
print(sheet.cell(row=6,column=1).value)
book.save("C:\\Users\\hp\\Documents\\Book1.xlsx")

print(sheet.max_row)
print(sheet.max_column)

dict={}
for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i,column=1).value == "t2":
        for j in range(2,sheet.max_column+1):
            print(sheet.cell(row=i,column=j).value,end=" ")
    print(end="\n")

